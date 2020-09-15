from openpyxl import load_workbook
from bs4 import BeautifulSoup
import random
import requests
import json
import os


def main():
    create_Lotto_json()
    Lotto(1, True)
    # temp()
    

# 로또 번호 추출 Main
def Lotto(edge = 3, edge_result = False, end_nums = [(1, 1), (7, 1)]):    
    print(edge)
    old_Lotto_numbers = get_list_LottoNums()
    nums_Lotto = []
    while len(nums_Lotto) < 5:
        num_Lotto = get_Lotto()
        
        if num_Lotto not in old_Lotto_numbers:
            # 모서리 패턴 체크 확인            
            if edge_result:
                if not edge_num(num_Lotto, edge):
                    continue

            nums_Lotto.append(num_Lotto)                          
            # if 끝수 패턴 Checked = True:
                # if not edn_num(num_Lotto, end_nums):
                    # continue
            
            # nums_Lotto.append(num_Lotto)
                    
            # if edge_num(num_Lotto, edge) and end_num(num_Lotto, end_nums):
                #and 끝수 :               

    return nums_Lotto
    # print(nums_Lotto)

# 임의의 숫자 6개 추출
def get_Lotto():
    예상수 = list(range(1, 46))#[10, 12, 13, 16, 17, 18, 20, 21, 22, 30, 33, 34, 35, 36, 37, 43, 45]
    약한수 = list(range(1, 46))
    고정수 = list(range(1, 46))#[10, 12, 13, 16, 17, 18, 20, 21, 22, 30, 33, 34, 35, 36, 37, 43, 45]
    확률 =  [0,1,2,3,4,5,6,7,8,9]
    예 = 0
    약 = 0
    고 = 0

    로또 = []

    for i in range(1, 7):
        if i < 2:
            if random.choice(확률) < 5:
                random.shuffle(고정수)
                num = random.choice(고정수)
                고정수.remove(num)
                예상수.remove(num)
                로또.append(num)
                고 += 1
            else:
                pro = random.choice(확률)            
                if pro < 1:
                    # print(pro)
                    random.shuffle(약한수)
                    num = random.choice(약한수)
                    약한수.remove(num)
                    로또.append(num)
                    약 += 1
                else:
                    random.shuffle(예상수)
                    num = random.choice(예상수)
                    예상수.remove(num)
                    로또.append(num)
                    예 += 1  
        else:
            pro = random.choice(확률)            
            if pro < 1:
                # print(pro)
                random.shuffle(약한수)
                num = random.choice(약한수)
                약한수.remove(num)
                로또.append(num)
                약 += 1
            else:
                random.shuffle(예상수)
                num = random.choice(예상수)
                예상수.remove(num)
                로또.append(num)  
                예 +=1          
    로또.sort()
    return 로또


# 끝수 파악
# return type이 true / false
# end_nums는 [ ( , ) ]로 전달받음
def end_num(numbers, end_nums):
    # 딕셔너리 제작
    end_num_dict = {}
    result = True
    for i in range(0, 10):
        end_num_dict[i] = 0
    
    for num in numbers:
        num = num % 10
        end_num_dict[num] += 1
    # print(numbers)
    for end_num in end_nums:
        if not end_num_dict[end_num[0]] >= end_num[1]:
            return False
    
    return result

# 끝수 내역 출력
def end_num_print(numbers):
    # 딕셔너리 제작
    end_num_dict = {}
    for i in range(0, 10):
        end_num_dict[i] = 0
    
    for num in numbers:
        num = num % 10
        end_num_dict[num] += 1

    end_num_dict = {key : val for key, val in end_num_dict.items() if val != 0}
    
    return end_num_dict
    
    # 끝수 확인 후 딕셔너리에 포함


# 모서리 확률
def edge_num(numbers, edge):
    edge_numbers = [1, 2, 8, 9, 6, 7, 13, 14, 29, 30, 36, 37, 34, 35, 41, 42]
    count = 0
    for num in numbers:
        if num in(edge_numbers):
            count += 1
    result = True if count > 0 and count <= edge else False
    return result

# 모서리 개수 출력    
def edge_num_print(numbers):
    edge_numbers = [1, 2, 8, 9, 6, 7, 13, 14, 29, 30, 36, 37, 34, 35, 41, 42]
    count = 0
    for num in numbers:
        if num*1.in(edge_numbers):
            count += 1    
    return count
    

# 이웃수 확률
def around_num(numbers):
    around_list = []
    # [3, 9, 13, 19, 23, 43]
    for num in numbers:                
        if not num-1.in(around_list):            
            around_list.append(num-1)
        if not num+1.in(around_list):
            around_list.append(num+1)
    return around_list


#  json을 만드는것(from 동행복권 parsing)
def create_Lotto_json():
    # Current Lotto rotation Load
    req_url = "https://dhlottery.co.kr/common.do?method=main"
    req_lotto = requests.get(req_url)
    soup = BeautifulSoup(req_lotto.content, 'html.parser')
    cur_lotto_num = soup.select("#lottoDrwNo")[0].get_text()

    # Create json file
    if os.path.isfile("./Lotto.json"):
        with open('./Lotto.json', 'r') as f:
            lotto_nums = json.load(f)
        if not cur_lotto_num in lotto_nums.keys():
            # print(cur_lotto_num)

            last_lotto_num_from_json = int(list(lotto_nums.keys())[0])
            
            for i in range( (int(cur_lotto_num) ), last_lotto_num_from_json, -1):
                # print(i)
                url = "https://www.dhlottery.co.kr/common.do?method=getLottoNumber&drwNo=" + str(i)
                req_lotto = requests.get(url)
                lottoNo = req_lotto.json()

                lotto_nums[int(i)] = [lottoNo['drwtNo1'],lottoNo['drwtNo2'],lottoNo['drwtNo3'],lottoNo['drwtNo4'],
                            lottoNo['drwtNo5'],lottoNo['drwtNo6'] ] # lottoNo['bnusNo']

            lotto_nums = dict(sorted(lotto_nums.items(), key=lambda k: int(k[0]), reverse=True))
            file_path = "./Lotto.json"
            with open(file_path, 'w') as outfile:
                json.dump(lotto_nums, outfile)
            
        else:
            print("존재")
    else:
        lotto_nums = {}

        for i in range(int(cur_lotto_num), 0, -1):
            # print(i)

            req_url = "https://www.dhlottery.co.kr/common.do?method=getLottoNumber&drwNo=" + str(i)
            req_lotto = requests.get(req_url)
            lottoNo = req_lotto.json()

            lotto_nums[int(i)] = [lottoNo['drwtNo1'],lottoNo['drwtNo2'],lottoNo['drwtNo3'],lottoNo['drwtNo4'],
                            lottoNo['drwtNo5'],lottoNo['drwtNo6'] ] # lottoNo['bnusNo']

        # print(lotto_nums)
        file_path = "./Lotto.json"
        
        with open(file_path, 'w') as outfile:
            json.dump(lotto_nums, outfile)


# 제작한 json파일을 List로 변환
def get_list_LottoNums():
    file_path = "./Lotto.json"
    with open(file_path, 'r') as f:
        lotto_nums = json.load(f)
    
    old_lotto_nums = list(val for val in lotto_nums.values())
    return old_lotto_nums

# 제작한 json 파일을 Dict로 변환
def get_dict_LottoNums():
    file_path = "./Lotto.json"
    with open(file_path, 'r') as f:
        lotto_nums = json.load(f)
    return lotto_nums




# deprecated
# 로또 번호 추출 함수
'''
def get_Lotto_numbers(episode):
    results = []
    file_name = "./Lotto.xlsx"
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.active
    for i in range(1, episode):
        result = []
        #result.append(ws['b'+str(i + 3)].value)
        result.append(ws['n'+str(i + 3)].value)
        result.append(ws['o'+str(i + 3)].value)
        result.append(ws['p'+str(i + 3)].value)
        result.append(ws['q'+str(i + 3)].value)
        result.append(ws['r'+str(i + 3)].value)
        result.append(ws['s'+str(i + 3)].value)
        results.append(result)
    
    #print(results)
    return results
'''


def temp():
    round = 928
    lotto_list = get_list_LottoNums()
    # print(lotto_list[-1:5:-1])
    for list in lotto_list[0:10]:
        print(round, "회차 끝수 정보 : ", end_num_print(list))
        print(round, "회차 모서리 개수 : ", edge_num_print(list))
        print(round, "회차 이웃수 개수 : ", around_num(list))
        round -= 1


if __name__ == "__main__":
    main()